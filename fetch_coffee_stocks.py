"""
ICE Coffee Certified Stock Report - Daily Fetcher
抓取 "TOTAL BAGS CERTIFIED" 中 "Total in Bags" 的 TOTAL 列数值
数据来源: https://www.ice.com/publicdocs/futures_us_reports/coffee/coffee_cert_stock_YYYYMMDD.xls
"""

import os
import sys
import datetime
import time
import tempfile
import requests
import xlrd
import openpyxl

# === 配置 ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "coffee_stocks_history.xlsx")
URL_TEMPLATE = "https://www.ice.com/publicdocs/futures_us_reports/coffee/coffee_cert_stock_{date}.xls"
MAX_RETRIES = 3
RETRY_DELAY = 10  # seconds

# === 日志 ===
def log(msg):
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {msg}")

# === 下载文件 ===
def download_report(date_str):
    """下载指定日期的 .xls 报告到临时目录，返回文件路径或 None"""
    url = URL_TEMPLATE.format(date=date_str)
    filename = f"coffee_cert_stock_{date_str}.xls"
    filepath = os.path.join(tempfile.gettempdir(), filename)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            log(f"下载: {url} (尝试 {attempt}/{MAX_RETRIES})")
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            }
            resp = requests.get(url, headers=headers, timeout=30)
            if resp.status_code == 200:
                with open(filepath, "wb") as f:
                    f.write(resp.content)
                log(f"下载成功: {filepath} ({len(resp.content)} bytes)")
                return filepath
            else:
                log(f"HTTP {resp.status_code}: {url}")
        except Exception as e:
            log(f"下载异常: {e}")

        if attempt < MAX_RETRIES:
            time.sleep(RETRY_DELAY)

    return None

# === 解析报告 ===
def parse_report(filepath):
    """
    解析 .xls 报告，返回 (report_date, total_bags) 或 (None, None)
    """
    # 从文件名提取日期作为 fallback
    basename = os.path.basename(filepath)
    fallback_date = None
    # coffee_cert_stock_20260615.xls → 提取 20260615
    import re
    m = re.search(r'(\d{8})', basename)
    if m:
        try:
            fallback_date = datetime.datetime.strptime(m.group(1), "%Y%m%d").date()
        except ValueError:
            pass

    try:
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        nrows = ws.nrows
        ncols = ws.ncols

        # 读取第3行(index 2)获取报告日期 (As of: Jun 15, 2026  1:31:49PM)
        report_date = None
        if nrows > 2:
            row1_text = str(ws.cell_value(2, 0)).strip()
            if "As of:" in row1_text:
                # 提取日期部分: "As of: Jun 15, 2026  1:31:49PM"
                date_part = row1_text.replace("As of:", "").strip()
                # 解析 "Jun 15, 2026"
                try:
                    # 取逗号前的月份日期和逗号后的年份
                    parts = date_part.split(",")
                    if len(parts) >= 2:
                        month_day = parts[0].strip()  # "Jun 15"
                        year = parts[1].strip()[:4]  # "2026"
                        month_str, day_str = month_day.split(" ")
                        month_map = {
                            "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
                            "May": 5, "Jun": 6, "Jul": 7, "Aug": 8,
                            "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12
                        }
                        month = month_map.get(month_str[:3], 1)
                        day = int(day_str)
                        report_date = datetime.date(int(year), month, day)
                except Exception:
                    pass

        # 扫描找到 "TOTAL BAGS CERTIFIED" 区域
        in_section = False
        for row_idx in range(nrows):
            cell_val = str(ws.cell_value(row_idx, 0)).strip().upper()
            if "TOTAL BAGS CERTIFIED" in cell_val:
                in_section = True
                continue

            if in_section:
                # 找到 "Total in Bags" 行
                first_cell = str(ws.cell_value(row_idx, 0)).strip()
                if first_cell.lower() == "total in bags":
                    # TOTAL 列是最后一列 (第9列，0-based)
                    total_val = ws.cell_value(row_idx, ncols - 1)
                    if isinstance(total_val, float):
                        total_bags = int(total_val)
                    else:
                        # 可能是字符串如 "397,242"
                        total_bags = int(str(total_val).replace(",", ""))
                    if report_date is None:
                        report_date = fallback_date
                    log(f"解析成功: date={report_date}, total_bags={total_bags}")
                    return report_date, total_bags

                # 如果遇到空行或下一个 section，退出
                if first_cell == "" and row_idx > 25:
                    break

        log("未找到 'Total in Bags' 行")
        return None, None

    except Exception as e:
        log(f"解析失败: {e}")
        return None, None

# === 更新历史 Excel ===
def update_history(date_val, bags_val):
    """
    将日期和数值追加到 OUTPUT_FILE，如果当天已有数据则更新
    """
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Coffee Stocks"
        ws.append(["Date", "Total Bags Certified", "Change"])
        # 设置列宽
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 14

    # 检查是否已有该日期的记录
    date_str = date_val.strftime("%Y-%m-%d") if date_val else "Unknown"
    updated = False
    for row in range(2, ws.max_row + 1):
        existing_date = ws.cell(row=row, column=1).value
        if isinstance(existing_date, datetime.datetime):
            existing_date = existing_date.date()
        if str(existing_date) == date_str:
            # 更新现有行
            ws.cell(row=row, column=2, value=bags_val)
            # 重新计算 Change
            if row > 2:
                prev_val = ws.cell(row=row - 1, column=2).value
                if prev_val is not None and isinstance(prev_val, (int, float)):
                    ws.cell(row=row, column=3, value=bags_val - prev_val)
            updated = True
            break

    if not updated:
        ws.append([date_str, bags_val, None])
        new_row = ws.max_row
        if new_row > 2:
            prev_val = ws.cell(row=new_row - 1, column=2).value
            if prev_val is not None and isinstance(prev_val, (int, float)):
                ws.cell(row=new_row, column=3, value=bags_val - prev_val)

    wb.save(OUTPUT_FILE)
    log(f"已更新历史文件: {OUTPUT_FILE} | {date_str}: {bags_val:,}")

# === 主流程 ===
def main():
    # 尝试今天、昨天、前天（应对周末/节假日）
    today = datetime.date.today()
    dates_to_try = [today - datetime.timedelta(days=i) for i in range(5)]

    for attempt_date in dates_to_try:
        date_str = attempt_date.strftime("%Y%m%d")
        filepath = download_report(date_str)
        if filepath is None:
            continue

        report_date, total_bags = parse_report(filepath)

        # 解析完立即清理临时 .xls 文件
        try:
            os.remove(filepath)
        except Exception:
            pass

        if report_date is not None and total_bags is not None:
            update_history(report_date, total_bags)
            return 0

    log("所有日期尝试均失败，请检查网络或 URL 是否仍然有效。")
    return 1

if __name__ == "__main__":
    sys.exit(main())
