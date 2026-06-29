"""
美国C型咖啡期货历史数据抓取脚本（修复版）
策略：yfinance（主，GitHub Actions US runner 可用）→ investing.com（备用）→ stooq（兜底）

修复点：
1. extract_from_next_data() 空正则 bug → 修复为正确的 __NEXT_DATA__ JSON 提取
2. yfinance 提升为主数据源，investing.com 降为备选
3. 新增 stooq.com 作为第三数据源
"""

import json
import re
import os
import sys
from datetime import datetime

# --- 配置 ---
INVESTING_URL = "https://cn.investing.com/commodities/us-coffee-c-historical-data"
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "coffee_price_history.xlsx")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}


# ============ 数据源 1: yfinance（主） ============

def fetch_from_yfinance():
    """从 Yahoo Finance 获取 KC=F 数据，GitHub Actions US runner 直接可用"""
    try:
        import yfinance as yf
        import pandas as pd
    except ImportError as e:
        print(f"[WARN] yfinance 不可用: {e}")
        return None

    print("[INFO] [yfinance] 正在获取 KC=F ...")
    try:
        ticker = yf.Ticker("KC=F")
        df = ticker.history(period="1mo")

        if df.empty:
            print("[WARN] [yfinance] 返回空数据")
            return None

        df = df.reset_index()
        raw_records = []
        prev_close = None
        for _, row in df.iterrows():
            date_str = row["Date"].strftime("%Y-%m-%d")
            close_val = float(row["Close"])
            open_val = float(row["Open"])
            high_val = float(row["High"])
            low_val = float(row["Low"])
            volume_val = int(row["Volume"]) if pd.notna(row["Volume"]) else 0

            if prev_close and prev_close != 0:
                change_pct = ((close_val / prev_close) - 1) * 100
            else:
                change_pct = 0

            raw_records.append({
                "rowDate": date_str,
                "last_close": f"{close_val:.2f}",
                "last_open": f"{open_val:.2f}",
                "last_max": f"{high_val:.2f}",
                "last_min": f"{low_val:.2f}",
                "volume": str(volume_val),
                "change_percent": f"{change_pct:.2f}",
            })
            prev_close = close_val

        # yfinance 返回降序，反转为升序
        raw_records.reverse()
        print(f"[INFO] [yfinance] 获取到 {len(raw_records)} 条记录")
        return raw_records

    except Exception as e:
        print(f"[WARN] [yfinance] 异常: {e}")
        return None


# ============ 数据源 2: investing.com（备用） ============

def fetch_investing_page():
    """获取 investing.com 页面 HTML"""
    # 方案1: curl_cffi 模拟 Chrome TLS 指纹
    try:
        from curl_cffi import requests as curl_requests
        print(f"[INFO] [investing] 正在请求页面 (curl_cffi): {INVESTING_URL}")
        session = curl_requests.Session()
        resp = session.get(INVESTING_URL, headers=HEADERS, impersonate="chrome", timeout=30)
        if resp.status_code == 200:
            print(f"[INFO] [investing] 页面获取成功，长度: {len(resp.text)} 字节")
            return resp.text
        print(f"[WARN] [investing/curl_cffi] 状态码: {resp.status_code}")
    except Exception as e:
        print(f"[WARN] [investing/curl_cffi] 失败: {e}")

    # 方案2: 降级到 requests
    try:
        import requests
    except ImportError:
        print("[ERROR] 缺少 requests 库")
        return None

    print(f"[INFO] [investing] 正在请求页面 (requests): {INVESTING_URL}")
    try:
        resp = requests.get(INVESTING_URL, headers=HEADERS, timeout=30)
        if resp.status_code != 200:
            print(f"[WARN] [investing/requests] 状态码: {resp.status_code}")
            return None
        print(f"[INFO] [investing] 页面获取成功，长度: {len(resp.text)} 字节")
        return resp.text
    except Exception as e:
        print(f"[WARN] [investing/requests] 异常: {e}")
        return None


def extract_from_next_data(html):
    """
    【修复】从 __NEXT_DATA__ JSON 中提取历史数据
    原 bug: re.search(r'', html, re.DOTALL) 空正则永远不匹配
    修复：正确的 Next.js SSR JSON 提取
    """
    # 匹配 <script id="__NEXT_DATA__" type="application/json"> ... </script>
    match = re.search(
        r'<script\s+id="__NEXT_DATA__"\s+type="application/json"\s*>(.*?)</script>',
        html,
        re.DOTALL,
    )
    if not match:
        print("[WARN] [investing] 未找到 __NEXT_DATA__")
        return None

    try:
        data = json.loads(match.group(1))
    except json.JSONDecodeError as e:
        print(f"[WARN] [investing] __NEXT_DATA__ JSON 解析失败: {e}")
        return None

    # 遍历可能的路径提取历史数据
    try:
        store = data["props"]["pageProps"]["state"]["historicalDataStore"]
        records = store["historicalData"]["data"]
        date_range = store.get("dateRange", {})
        print(f"[INFO] [investing] __NEXT_DATA__ 提取到 {len(records)} 条记录")
        print(f"[INFO] [investing] 日期范围: {date_range.get('startDate', '?')} ~ {date_range.get('endDate', '?')}")
        return records
    except (KeyError, TypeError) as e:
        print(f"[WARN] [investing] __NEXT_DATA__ 结构解析失败: {e}")
        return None


def extract_from_html_tables(html):
    """备用方案: 从 HTML 表格中提取历史数据"""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        print("[ERROR] 缺少 beautifulsoup4 库")
        return None

    soup = BeautifulSoup(html, "lxml")
    tables = soup.find_all("table")

    for table in tables:
        headers = [th.get_text(strip=True) for th in table.find_all("th")]
        if "日期" in headers and "收盘" in headers:
            records = []
            rows = table.find_all("tr")[1:]
            for row in rows:
                cells = row.find_all("td")
                if len(cells) >= 7:
                    records.append({
                        "rowDate": cells[0].get_text(strip=True),
                        "last_close": cells[1].get_text(strip=True),
                        "last_open": cells[2].get_text(strip=True),
                        "last_max": cells[3].get_text(strip=True),
                        "last_min": cells[4].get_text(strip=True),
                        "volume": cells[5].get_text(strip=True) if len(cells) > 5 else "",
                        "change_percent": cells[6].get_text(strip=True) if len(cells) > 6 else "",
                    })
            if records:
                print(f"[INFO] [investing] HTML 表格提取到 {len(records)} 条记录")
                return records
    return None


def fetch_from_investing():
    """investing.com 完整流程"""
    html = fetch_investing_page()
    if not html:
        return None

    records = extract_from_next_data(html)
    if not records:
        print("[WARN] [investing] __NEXT_DATA__ 提取失败，尝试 HTML 表格...")
        records = extract_from_html_tables(html)

    return records


# ============ 数据源 3: stooq.com（兜底） ============

def fetch_from_stooq():
    """从 stooq.com 获取 KC 期货数据（纯 CSV，无 Cloudflare）"""
    import urllib.request
    import csv
    import io

    # stooq 咖啡期货代码: KCH26 — 实时主力合约
    stooq_url = "https://stooq.com/q/d/l/?s=kc.f&i=d"
    print(f"[INFO] [stooq] 正在请求: {stooq_url}")

    try:
        req = urllib.request.Request(stooq_url, headers={"User-Agent": HEADERS["User-Agent"]})
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode("utf-8")
    except Exception as e:
        print(f"[WARN] [stooq] 请求失败: {e}")
        return None

    # stooq CSV 格式: Date,Open,High,Low,Close,Volume
    reader = csv.DictReader(io.StringIO(raw))
    raw_records = []
    prev_close = None

    for row in reader:
        date_str = row.get("Date", "")
        if not date_str:
            continue
        try:
            close_val = float(row["Close"])
            open_val = float(row["Open"])
            high_val = float(row["High"])
            low_val = float(row["Low"])
            volume_val = int(float(row.get("Volume", 0))) if row.get("Volume") else 0
        except (ValueError, TypeError):
            continue

        if prev_close and prev_close != 0:
            change_pct = ((close_val / prev_close) - 1) * 100
        else:
            change_pct = 0

        raw_records.append({
            "rowDate": date_str,
            "last_close": f"{close_val:.2f}",
            "last_open": f"{open_val:.2f}",
            "last_max": f"{high_val:.2f}",
            "last_min": f"{low_val:.2f}",
            "volume": str(volume_val),
            "change_percent": f"{change_pct:.2f}",
        })
        prev_close = close_val

    # stooq 返回降序，反转为升序
    raw_records.reverse()
    if raw_records:
        print(f"[INFO] [stooq] 获取到 {len(raw_records)} 条记录")
    else:
        print("[WARN] [stooq] 返回空数据")
    return raw_records or None


# ============ 数据标准化 ============

def normalize_records(raw_records):
    """将原始记录转换为统一格式"""
    result = []
    for rec in raw_records:
        item = {
            "日期": rec.get("rowDate", ""),
            "收盘价": rec.get("last_close", ""),
            "开盘价": rec.get("last_open", ""),
            "最高价": rec.get("last_max", ""),
            "最低价": rec.get("last_min", ""),
            "交易量": rec.get("volume", ""),
            "涨跌幅": "",
        }
        pct = rec.get("change_percent", "") or rec.get("change_percent", "")
        if pct:
            try:
                item["涨跌幅"] = f"{float(pct):+.2f}%"
            except ValueError:
                item["涨跌幅"] = pct
        result.append(item)
    return result


# ============ 增量合并与保存 ============

def load_existing_records():
    """从已有 Excel 加载历史记录"""
    if not os.path.exists(OUTPUT_FILE):
        print("[INFO] 未找到已有数据文件，将创建新文件")
        return []

    try:
        import pandas as pd
        df = pd.read_excel(OUTPUT_FILE, header=3)
        records = []
        for _, row in df.iterrows():
            if pd.notna(row.get("日期")):
                records.append({
                    "日期": str(row["日期"])[:10],
                    "收盘价": str(row.get("收盘价", "")),
                    "开盘价": str(row.get("开盘价", "")),
                    "最高价": str(row.get("最高价", "")),
                    "最低价": str(row.get("最低价", "")),
                    "交易量": str(row.get("交易量", "")),
                    "涨跌幅": str(row.get("涨跌幅", "")),
                })
        if records:
            print(f"[INFO] 从已有文件加载 {len(records)} 条历史记录")
        return records
    except Exception as e:
        print(f"[WARN] 读取已有文件失败: {e}，将创建新文件")
        return []


def save_to_excel(records):
    """保存数据到 Excel 文件"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print("[ERROR] 缺少 openpyxl 库")
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "咖啡期货历史数据"

    ws.merge_cells("A1:G1")
    ws["A1"] = "美国C型咖啡期货 (KC) — 历史数据"
    ws["A1"].font = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = f"数据来源: Yahoo Finance (KC=F)  |  抓取时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(name="Microsoft YaHei", size=9, color="808080")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 22

    headers = ["日期", "收盘价", "开盘价", "最高价", "最低价", "交易量", "涨跌幅"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[4].height = 24

    red_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    green_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")

    for row_idx, rec in enumerate(records, 5):
        for col_idx, key in enumerate(headers, 1):
            value = rec.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Microsoft YaHei", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            if key == "涨跌幅" and value:
                try:
                    if float(value.replace("%", "").replace("+", "")) >= 0:
                        cell.fill = green_fill
                    else:
                        cell.fill = red_fill
                except ValueError:
                    pass
        ws.row_dimensions[row_idx].height = 22

    col_widths = [18, 12, 12, 12, 12, 12, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:G{4 + len(records)}"
    wb.save(OUTPUT_FILE)
    print(f"[INFO] Excel 文件已保存: {OUTPUT_FILE}")


# ============ 主流程 ============

def main():
    print("=" * 60)
    print("  美国C型咖啡期货 (KC) 历史数据抓取工具（修复版）")
    print("=" * 60)

    # 数据源优先级: yfinance → investing.com → stooq
    raw_records = None
    source_name = ""

    # 1. yfinance（主数据源，GitHub Actions US runner 原生可用）
    raw_records = fetch_from_yfinance()
    if raw_records:
        source_name = "Yahoo Finance (KC=F)"

    # 2. investing.com（备用，需绕过 Cloudflare）
    if not raw_records:
        print("[INFO] yfinance 不可用，尝试 investing.com ...")
        raw_records = fetch_from_investing()
        if raw_records:
            source_name = "Investing.com"

    # 3. stooq（兜底）
    if not raw_records:
        print("[INFO] investing.com 不可用，尝试 stooq ...")
        raw_records = fetch_from_stooq()
        if raw_records:
            source_name = "stooq.com"

    if not raw_records:
        print("[ERROR] 所有数据源均失败，请稍后重试。")
        sys.exit(1)

    print(f"[INFO] 成功从 {source_name} 获取数据")

    records = normalize_records(raw_records)

    # 去重
    seen = set()
    deduped = []
    for rec in records:
        date_key = rec["日期"]
        if date_key not in seen:
            seen.add(date_key)
            deduped.append(rec)
    if len(deduped) < len(records):
        print(f"[INFO] 新数据去重: {len(records)} -> {len(deduped)} 条")

    # 升序排列
    deduped.reverse()

    # 加载已有历史数据并合并
    old_records = load_existing_records()
    merged = {}
    for rec in old_records:
        merged[rec["日期"]] = rec
    new_count = 0
    for rec in deduped:
        date_key = rec["日期"]
        if date_key not in merged:
            new_count += 1
        merged[date_key] = rec

    all_records = sorted(merged.values(), key=lambda x: x["日期"])
    print(f"[INFO] 合并完成：历史 {len(old_records)} 条 + 新增 {new_count} 条 = 共 {len(all_records)} 条")

    save_to_excel(all_records)

    print(f"\n[预览] 共 {len(all_records)} 条记录:")
    print("-" * 80)
    for rec in all_records[-10:]:
        print(
            f"  {rec['日期']:16s} | "
            f"收盘:{rec['收盘价']:>8s} | "
            f"开盘:{rec['开盘价']:>8s} | "
            f"最高:{rec['最高价']:>8s} | "
            f"最低:{rec['最低价']:>8s} | "
            f"量:{rec['交易量']:>8s} | "
            f"涨跌:{rec['涨跌幅']}"
        )
    if len(all_records) > 10:
        print(f"  ... (还有 {len(all_records) - 10} 条)")

    print(f"\n[DONE] 数据已保存到: {OUTPUT_FILE}")
    print(f"[INFO] 数据源: {source_name} | 本次新增 {new_count} 条 | 累计 {len(all_records)} 条")


if __name__ == "__main__":
    main()
