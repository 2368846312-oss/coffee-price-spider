import sys, os, io, re, time
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(SCRIPT_DIR, "咖啡豆每日价格.xlsx")

LOCATIONS_ORDER = [
    "全国", "云南", "云南:普洱", "云南:普洱:思茅区", "云南:普洱:景谷县",
    "云南:保山", "云南:保山:隆阳区", "云南:临沧", "云南:临沧:云县"
]

def create_driver():
    """自动选择可用浏览器：Edge(本地) -> Chrome(GitHub Actions)"""
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options as ChromeOptions

    # 尝试 Edge（本地 Windows）
    edge_path = "C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe"
    if os.path.isfile(edge_path):
        from selenium.webdriver.edge.options import Options
        opts = Options()
        opts.binary_location = edge_path
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_experimental_option("excludeSwitches", ["enable-logging"])
        return webdriver.Edge(options=opts)

    # 尝试 Chrome（GitHub Actions）
    from selenium.webdriver.chrome.options import Options
    opts = ChromeOptions()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_experimental_option("excludeSwitches", ["enable-logging"])
    return webdriver.Chrome(options=opts)

def extract_prices(page_text):
    lines = page_text.split("\n")
    national_avg = None
    for i, line in enumerate(lines):
        if "今日均价" in line:
            for j in range(i, min(i+5, len(lines))):
                m = re.search(r"(\d+\.?\d*)", lines[j])
                if m:
                    national_avg = float(m.group(1))
                    break
            break

    target_regions = [
        ("云南", lambda l: "云南" in l and not any(x in l for x in ["昆明", "保山", "普洱", "临沧", "大理", "红河", "版纳", "文山", "思茅", "隆阳", "景谷", "云县", "凤庆", "龙陵"])),
        ("云南:普洱", lambda l: "云南普洱" in l and "思茅" not in l and "景谷" not in l),
        ("云南:普洱:思茅区", lambda l: "思茅区" in l),
        ("云南:普洱:景谷县", lambda l: "景谷" in l),
        ("云南:保山", lambda l: "云南保山" in l and "隆阳" not in l),
        ("云南:保山:隆阳区", lambda l: "隆阳区" in l),
        ("云南:临沧", lambda l: "云南临沧" in l and "云县" not in l),
        ("云南:临沧:云县", lambda l: "云县" in l),
    ]

    prices = {}
    for i, line in enumerate(lines):
        pm = re.search(r"(\d+\.\d+)元/斤", line)
        if pm:
            price = float(pm.group(1))
            for j in range(max(0, i-4), i):
                prev = lines[j].strip()
                for name, checker in target_regions:
                    if checker(prev) and name not in prices:
                        prices[name] = price
                        break

    return national_avg, prices

def ensure_excel():
    if os.path.isfile(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "咖啡豆每日价格"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ws.append(["日期"] + LOCATIONS_ORDER)
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
        cell.border = tb
    ws.column_dimensions["A"].width = 14
    for i in range(2, len(LOCATIONS_ORDER) + 2):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 18
    wb.save(EXCEL_FILE)
    return wb, ws

def update_excel(national_avg, prices, today):
    wb, ws = ensure_excel()
    tb = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    existing = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        existing.append(str(row[0]) if row[0] else "")

    row_data = [today]
    for loc in LOCATIONS_ORDER:
        if loc == "全国":
            row_data.append(national_avg if national_avg else "#N/A")
        else:
            row_data.append(prices.get(loc, "#N/A"))

    if today in existing:
        for r_idx in range(2, ws.max_row + 1):
            if str(ws.cell(row=r_idx, column=1).value) == today:
                for c_idx, val in enumerate(row_data[1:], 2):
                    ws.cell(row=r_idx, column=c_idx, value=val)
                break
        print(f"  ✓ 更新 {today} 数据")
    else:
        ws.append(row_data)
        print(f"  ✓ 新增 {today} 数据")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(LOCATIONS_ORDER)+1):
        for cell in row:
            cell.border = tb
            cell.alignment = Alignment(horizontal="center")
    wb.save(EXCEL_FILE)
    print(f"  ✓ Excel 已保存")

def main():
    print("=" * 50)
    print("☕ 咖啡豆价格爬虫开始运行")
    print(f"   时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    try:
        driver = create_driver()
        print(f"   🌐 正在获取 https://www.ymt.com/hangqing/juhe-26732 ...")
        driver.get("https://www.ymt.com/hangqing/juhe-26732")
        time.sleep(4)
        page_text = driver.find_element("tag name", "body").text
        driver.quit()
        print(f"   ✅ 页面加载成功 ({len(page_text)} 字符)")
    except Exception as e:
        print(f"   ❌ 页面获取失败: {e}")
        return

    national_avg, prices = extract_prices(page_text)
    today = datetime.now().strftime("%Y-%m-%d")

    print(f"\n   📅 日期: {today}")
    if national_avg:
        print(f"   💰 全国均价: {national_avg} 元/斤")
    else:
        print(f"   💰 全国均价: 未获取到")

    for loc in LOCATIONS_ORDER:
        if loc == "全国":
            val = national_avg if national_avg else "#N/A"
        else:
            val = prices.get(loc, "#N/A")
        print(f"      {loc}: {val}")

    print(f"\n   💾 正在保存到 Excel...")
    update_excel(national_avg, prices, today)

    print(f"\n✅ 完成！Excel: {EXCEL_FILE}")
    print("=" * 50)

if __name__ == "__main__":
    main()

